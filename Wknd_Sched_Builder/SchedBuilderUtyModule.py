from SchedBuilderClasses2 import *
import openpyxl as pyxl
import pandas as pd
import numpy as np
import sqlite3
import functools
from copy import deepcopy


def debug(func):
    """Print the function signature and return value"""
    @functools.wraps(func)
    def wrapper_debug(*args, **kwargs):
        args_repr = [repr(a) for a in args]                      # 1
        kwargs_repr = [f"{k}={v!r}" for k, v in kwargs.items()]  # 2
        signature = ", ".join(args_repr + kwargs_repr)           # 3
        print(f"Calling {func.__name__}({signature})")
        value = func(*args, **kwargs)
        print(f"{func.__name__!r} returned {value!r}")           # 4
        return value
    return wrapper_debug



def addTBL(tblName,fields="",dTypes=None,data=None,addOn=False):
    """Create table if not already existing, optionally with data, optionally clearing out old data if present. Fields as list of strings. Datatypes as list of strings, one must be provided for each field. See sqlite3 docs for mroe info"""
    conn = sqlite3.connect('test17.db')
    c = conn.cursor()
    listedFields=''
    if fields=="": #If none given, make alphabetical
        fields=[chr(65+i) for i in range(len(data[0]))]        
    if dTypes==None: #Need not specify dtypes
        for f in fields:
            listedFields=listedFields+', '+ f
    else: #define data types at inception of table
        flds=list(zip(fields,dTypes))
        for pair in flds:
            listedFields=listedFields+', '+pair[0]+' '+pair[1] 
    listedFields='('+listedFields[2:]+''')''' #Add leading and closing bracket, remove naively added comma+space from leading field
    c.execute('''CREATE TABLE IF NOT EXISTS '''+tblName+listedFields) # Create table.
    if addOn==False: #Delete if not adding
        c.execute('''DELETE FROM '''+tblName)
    if (data is not None) and len(data)>0:
        stmnt='INSERT INTO '+tblName+' VALUES ('
        for i in range(len(fields)-1):
            stmnt=stmnt+'?,'#Add '?,' equal to num columns less 1
        stmnt=stmnt+'?)' #add closing ?), no final comma
        for subEntry in data:
            c.execute(stmnt, subEntry)
    conn.commit()

def isNumeric(n):
    try:
        n=int(n)
        return True
    except ValueError:
        try:
            n=float(n)
            return True
        except:
            return False


def viewTBL(tblName,fields=None,sortBy=None,filterOn=None,returnStatement=0):
    """return np array of table with optional select fields, filtered, sorted. Sort syntax=[(field1,asc/desc),(field2,asc/desc)...] Filter syntax=[(field1,value),(field2,value)...]"""
    conn = sqlite3.connect('test17.db')
    c = conn.cursor()
    stmnt='SELECT '
    if fields!=None: 
        flds=''
        for f in fields:
            flds=flds+', '+f
        stmnt=stmnt+flds[2:]+ ' FROM ' +tblName+' '
    else: stmnt=stmnt+'* FROM '+tblName+' ' #unspecified, select all
    if filterOn!=None:
        filt='WHERE '
        for f in filterOn:
            if isNumeric(f[1]): filt=filt+f[0]+' = '+ str(f[1])+' AND '
            else: filt=filt+str(f[0])+' = "'+ str(f[1])+'" AND '
        filt=filt[:-4] #Remove naively added final " and "
        stmnt=stmnt+filt
    if sortBy!=None:
        srt='ORDER BY '
        for s in sortBy:
            srt=srt+s[0]+' '+s[1]+', '
        srt=srt[:-2]
        stmnt=stmnt+srt
    stmnt=stmnt+';'
    if returnStatement==True: # Add option to print out the sql statement for troubleshooting
        return stmnt
    else:
        c.execute(stmnt)
        return [list(x) for x in c.fetchall()] #sqlite3 returns list of tuples.. want sublists for being editable

def FTbtRow(ws):
    """Returns the excel row number for the bottom row with data in FT employee sheet"""
    #1st find bottom row with data
    for i in range(5,400):
        ref="C"+str(i) #Referencing EEID column. Failure mode is EEID missing for someone. Thats a bigger problem than the code not working
        if ws[ref].internal_value==None: 
            #Condition met when end of data found.
            btmRow=i-1 #backtrack to last data
            break
    return btmRow

def getFTinfo(flNm):
    """Returns dataframe with FT employee info (seniority,crew,eeid,name,refusal hours to date, OT hrs worked this week given path to FT refusal sheet"""
    myWb=pyxl.load_workbook(flNm)
    ws=myWb['Hourly OT']
    btmRow=FTbtRow(ws)
    tab=[[x.internal_value for x in sublist] for sublist in ws['A5:I'+str(btmRow)]]
    #for rec in tab: #numeric values getting cast to string on import. cast back
    #    for i in [0,2]:
    #        rec[i]=int(rec[i])
    #    for i in [5,6,7]:
    #        rec[i]=float(rec[i])
    #Following to turn into dataframe
    #df_FTinfo=pd.DataFrame(tab)
    #df_FTinfo=df_FTinfo[[0,1,2,3,4,5,8]] #Pull out only required columns 
    #df_FTinfo.set_axis(['snrty', 'crew', 'eeid','last','first','yrRef','wkOT'], axis='columns', inplace=True)
    return tab


def FTendCol(ws):
    """Returns column # for last column of skills matrix in excel from FT refusal sheet"""
    for i in range(0,400):
        if ws['A1'].offset(0,i).value=='Start-up': 
            #Condition met when end of data found.
            endCol=i-1
            break
    return endCol

def getFTskills(flNm):
    """Returns a dataframe containing a table with 2 fields: eeid and job name, with one record for every job an ee is trained in"""
    myWb=pyxl.load_workbook(flNm)
    ws=myWb['Hourly OT']
    endCol=FTendCol(ws) #Get right limit for iteration through skills
    btmRow=FTbtRow(ws) #Get bottom limit for iteration through skills
    skills=[] #Initialize empty skills list
    for i in range(5,btmRow+1): #Data starts on row 5. +1 because range fn not inclusive
        eeid=ws['C'+str(i)].value
        for c in range(10,endCol+1):
            if ws['C'+str(i)].offset(0,c-2).value==1: #subtract 2 from c because the endCol is counted from col A, and we are offsetting from col C, whihc is 2 offset from A
                jobNm=ws['A2'].offset(0,c).value #if 1 indicates trained, pull job name from header row
                skills.append([eeid,jobNm]) #Add new record to skills table
    #idxs=np.array(skills)[:,0]
    #skills=pd.DataFrame(skills,idxs) #Convert to dataframe
    #skills.set_axis(['eeid','skill'], axis='columns', inplace=True)
    return skills

def TempbtRow(ws):
    """Returns the excel row number for the bottom row with data in Temp employee sheet"""
    #1st find bottom row with data
    for i in range(4,400):
        ref="C"+str(i) #Referencing EEID column. Failure mode is EEID missing for someone. Thats a bigger problem than the code not working
        if ws[ref].internal_value==None: 
            #Condition met when end of data found.
            btmRow=i-1 #backtrack to last data
            break
    return btmRow

def getTempinfo(flNm):
    """Returns dataframe with FT employee info (seniority,crew,eeid,name,refusal hours to date, OT hrs worked this week given path to FT refusal sheet"""
    myWb=pyxl.load_workbook(flNm)
    ws=myWb['Temp Refusal']
    btmRow=TempbtRow(ws)
    tab=[[x.internal_value for x in sublist] for sublist in ws['A4:I'+str(btmRow)]]
    #df_Tempinfo=pd.DataFrame(tab)
    #df_Tempinfo=df_Tempinfo[[0,1,2,3,4,5,8]] #Pull out only required columns 
    #df_Tempinfo.set_axis(['snrty', 'crew', 'eeid','last','first','yrRef','wkOT'], axis='columns', inplace=True)
    return tab

def TempendCol(ws):
    """Returns column # for last column of skills matrix in excel from FT refusal sheet"""
    for i in range(0,400):
        if ws['A2'].offset(0,i).value=='Start Up': 
            #Condition met when end of data found.
            endCol=i-1
            break
    return endCol

def getTempskills(flNm):
    """Returns a dataframe containing a table with 2 fields: eeid and job name, with one record for every job an ee is trained in"""
    myWb=pyxl.load_workbook(flNm)
    ws=myWb['Temp Refusal']
    endCol=TempendCol(ws) #Get right limit for iteration through skills
    btmRow=TempbtRow(ws) #Get bottom limit for iteration through skills
    skills=[] #Initialize empty skills list
    for i in range(4,btmRow+1): #Data starts on row 5. +1 because range fn not inclusive
        eeid=ws['C'+str(i)].value
        for c in range(11,endCol+1): #First skills column is 11 offset from col A
            if ws['C'+str(i)].offset(0,c-2).value==1: #subtract 2 from c because the endCol is counted from col A, and we are offsetting from col C, which is 2 offset from A
                jobNm=ws['A3'].offset(0,c).value #if 1 indicates trained, pull job name from header row
                skills.append([eeid,jobNm]) #Add new record to skills table
    #idxs=np.array(skills)[:,0]
    #skills=pd.DataFrame(skills,idxs) #Convert to dataframe
    #skills.set_axis(['eeid','skill'], axis='columns', inplace=True)
    return skills

def imptXlTbl(XlFl,ShtNm,TblNm):
    myWb=pyxl.load_workbook(XlFl) 
    ws=myWb[ShtNm]
    tab=ws.tables[TblNm] #Pull out table
    tab=[[x.value for x in sublist] for sublist in ws[tab.ref]] #Convert to list of lists (each sublist as row of excel table)
    return tab[1:] #Convert nested lists to array, dropping first row which is table headings

def imptPolltbl(XlFl,ShtNm,TblNm,tp=None):
    myWb=pyxl.load_workbook(XlFl) 
    ws=myWb[ShtNm]
    tab=ws.tables[TblNm] #Pull out table
    tab=[[x.value for x in sublist] for sublist in ws[tab.ref]] #Convert to list of lists (each sublist as row of excel table)
    tab=tab[1:] #Remove header column
    if tp=='FT': #Pull only FT's into FT table
        tab=[rec for rec in tab if rec[3]!=None] #Remove rows without refusal hours
        tab=[rec for rec in tab if rec[3]<10000] #Ft's id'd by less than 10k refusal hours
    if tp=='P': #Pull only probationaries into table
        tab=[rec for rec in tab if rec[3]!=None] #Remove rows without refusal hours
        tab=[rec for rec in tab if rec[3]>=10000] #Probationaries ID'd by 10K or more refusal hours
    return tab

def generateMasterPollTbl(pollDict):
    """Given a dictionary containing the polling tables for all crews, generates a master tbl in SQLlite for being able to filter on peoples availabilities, with '1' indicating interest, '0' no interest, and slot seq 1 starting at index 4"""
    mPollTbl=[]
    #the total list of all fields the table has is programmatically generated on these 3 lines
    flds=["eeid",'lastNm','firstNm','ytdRefHrs'] 
    flds.extend(['slot_'+str(i) for i in range(1,25)])#Note that there is one field for each slot seqID, 1 through 24, for filtering
    flds.append('Comment')
    for crewKey in pollDict:
        tbl=pollDict[crewKey] #Pull the crew specific OT polling table from dictionary
        for rec in tbl:
            cmnt=rec[16]#retrieve comment to tag on later
            slotwise_polling=list(rec[:4])
            for i in range(4,16):
                if rec[i] not in ('n','N',None) :
                    slotwise_polling.extend(['y','y']) #Add two entries because 1 entry in polling sheet applies to two slots
                else:
                    slotwise_polling.extend(['n','n'])
            slotwise_polling.append(cmnt)
            mPollTbl.append(slotwise_polling)
    addTBL('allPollData',fields=flds, data=mPollTbl,addOn=False)


def pullTbls(FtBook,TempBook,AssnBook,PollBook):  #Need to make volunteer shift data puller
    """Take flNm, return ftInfoTbl, ftSkillsMtx, tempInfoTbl, tempSkillsMtx, AssignmentsTbl, slot_Legend, JobTrnCrossRef, pollDict, All_Slots, senList.   Uses functions defined previously to return all required tables at once. Function of functions for final script"""
    a=getFTinfo(FtBook) #to sqlite
    b=getFTskills(FtBook) #to sqlite
    b=[[int(d[0]),d[1]] for d in b] #Cast EEid to numeric value
    c=getTempinfo(TempBook) #to sqlite
    d=getTempskills(TempBook) #to sqlite
    d=[[int(data[0]),data[1]] for data in d] #Cast EEid to numeric value
    e=imptXlTbl(AssnBook,'Assignment_List','Assn_List')
    f=imptXlTbl(AssnBook,'Slot_Legend','Slot_Legend')
    g=imptXlTbl(AssnBook,'Job_Training_Crossref','TrainAssnMtx') #to sqlite
    pollDict={} #Generate empty dictionary to store tables of people voluntary overtime
    for crew in ['Blue','Bud','Rock']:
        for eeType in ['FT','P','Temp']:
            if eeType=='Temp': #If type= Temp, proceed to build table
                keyNm='tbl_'+crew+eeType
                tbl=imptXlTbl(PollBook,'Sheet1',keyNm)
            else:
                keyNm='tbl_'+crew+'FT' #If type= FT OR Probationary, will be referring to FT table in excel, so hard code the string
                tbl=imptPolltbl(PollBook,'Sheet1',keyNm,tp=eeType)
            if eeType=='P': #keyNm was made "FT" instead of 'P' so need to manually enter the key when generating dictionry entry
                pollDict['tbl_'+crew+'P']=tbl
            else:
                pollDict[keyNm]=tbl
    pollDict['tbl_wFT']=imptPolltbl(PollBook,'Sheet1','tbl_wFT',tp='FT') #No nice loop to initialize the WWF crew tables in poll sheet
    pollDict['tbl_wP']=imptPolltbl(PollBook,'Sheet1','tbl_wFT',tp='P')
    pollDict['tbl_wT']=imptXlTbl(PollBook,'Sheet1','tbl_wT')
    h=imptXlTbl(AssnBook,'All_Slots','All_Slots')
    #Generate tables in sqlite
    addTBL("sklMtx",fields=["EEID","trnNm"],data=b,addOn=False) #Overwrite all training data and populate FT ops, then append temps for a master table
    addTBL("sklMtx",fields=["EEID","trnNm"],data=d,addOn=True)
    addTBL("xRef",fields=["dispNm","trnNm"],data=g,addOn=False) #Skill name cross ref table for fcn dispToTrn to work
    addTBL("FTinfo",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=a,addOn=False)
    addTBL("TempInfo",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=c,addOn=False)
    #Generate a master seniority table.. following replaces hire date with integers for temps
    senHiLoTemps=viewTBL('TempInfo',sortBy=[('sen','ASC')]) #First retrieve list of temps, most senior to least
    i=100000 #Start new seniority number at arbitrarily high value not to interfere with full timer
    for row in senHiLoTemps:
        row[0]=i
        i+=1
    #Overwrite/make new master sen ref table. Then append the Temp data with integerized values
    addTBL("senRef",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=a,addOn=False)
    addTBL("senRef",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=senHiLoTemps,addOn=True)
    senList=viewTBL('senRef',sortBy=[('sen','ASC')])
    return a,b,c,d,e,f,g,pollDict,h,senList

def dispToTrn(dispNm):
    """Returns the trnNm associated with Display name for a given job. assumes popualted sqlite table 'xRef' with dispNm/trnNm pairs"""
    q=viewTBL('xRef',fields=['dispNm','trnNm'],filterOn=[('dispNm',dispNm)])
    if len(q)==0:
        return "Custom func error 'dispToTrn' no entry found in xRef with dispNm="+str(dispNm)
    return q[0][1]

def trnToDisp(trnNm):
    """Returns the trnNm associated with Display name for a given job. assumes popualted sqlite table 'xRef' with dispNm/trnNm pairs"""
    q=viewTBL('xRef',fields=['dispNm','trnNm'],filterOn=[('trnNm',trnNm)])
    if len(q)==0:
        return "Custom func error 'trnToDisp' no entry found in xRef with trnNm="+str(trnNm)
    return [e[0] for e in q] #If multiple DispNms for one train name (e.g. L4 Packer -> Packer, Candling)  or Bottle Supply -> etc.
                            #Then return list of all dispNms

def sklChk(eeid,dispNm):
    """Returns True/False if eeid is trained on job with display name or not. Requires skills matrix named sklMtx in sqlite"""
    trnNm=dispToTrn(dispNm)
    if len(viewTBL('sklMtx',filterOn=[('EEID',eeid),('trnNm',trnNm)]))==0:
        return False
    else:
        return True


def makeEEdict(ftInfoTbl,tempInfoTbl,wkHrs=40,tp='id'):
    eeDict={}
    for dtaTbl in [ftInfoTbl,tempInfoTbl]:
        for row in dtaTbl:
            # if row[1].lower().strip() in ['wwf','bud','blue','rock','silver','gold','student']: #Omit people not in packaging, or off, vacation etc
            if row[2] not in list(eeDict.keys()): #Double check ee hasn't already been generated... why Cory would include an ee on temp table with crew reading 'fulltime' is beyond me but there you go
                if tp=='id':
                    eeSkills=viewTBL('sklMtx',['trnNm'],filterOn=[('EEID',row[2])])
                    eeSkills=[trnToDisp(nm[0]) for nm in eeSkills] #Gather display names for skills trained on, reducing lists within list to spread elements
                    sk=[] #Create empty to accumulate all skills present within sublists of eeSkills
                    for s in eeSkills:
                        sk.extend(s)
                    sen=viewTBL('senRef',fields=['sen'],filterOn=[('id',str(row[2]))])[0][0]
                    anEE=ee(sen,row[1].lower().strip(),int(row[2]),row[3],row[4],row[5],row[8]+wkHrs,skills=sk) #Pull info from Refusals sheet
                    eeDict[anEE.eeID]=anEE
                elif tp=='nm':
                    #Just need peoples names and EEID's linked from this.. enter dummy data for seniority, training, etc
                    if dtaTbl==tempInfoTbl:sen=100000
                    else: sen=1
                    anEE=ee(sen,row[1].lower().strip(),int(row[2]),row[3],row[4],row[5],row[8]+wkHrs,skills=[])
                    eeDict[anEE.dispNm().lower().replace(' ','-')]=anEE
    return eeDict

def makeSlots(eeDict,AllSlots):
    openSlots={} #Open here meaning unassigned.. Will be required when it comes time to force
    for row in AllSlots:
        if row[6]==1: #Check that the slot generation record is labelled as 'active'
            for i in range(row[0],row[1]+1): #Generate a slot for each index over the range indicated... add 1 because python Range fn not inclusive of end point
                sl=Slot(i, row[2],dispToTrn(row[2]))
                #Determine how many eligible volunteers for this slot
                elig=[] #To track how many people trained
                for rec in viewTBL('allPollData',filterOn=[('slot_'+str(sl.seqID),'y')]): # iterate through results (employee info's) of query on who said yes to working at the time of this slot
                    if sl.dispNm in eeDict[rec[0]].skills: elig.append(rec[0]) #Append EEID to list 'elig' if the ee is trained on the job
                sl.eligVol=elig # TTake len() to see number of eligible volunteers for the slot.
                openSlots[str(sl.seqID)+'_'+str(sl.dispNm)]=sl #Enter it into the dictionary
    return openSlots

def preProcessData(Acrew,wkHrs,FtBook,TempBook,AssnBook,PollBook,pNT=False,assnWWF=False,pVol=True,xtraDays=None,maxI=100):
    """A function to take input data and generate all necessary tables and objects in memory to carry out algorithm. Return Schedule object containing all workSlot objects, and dictioanry fo all employee objects"""
    ftInfoTbl, ftSkillsMtx, tempInfoTbl, tempSkillsMtx, AssignmentsTbl, slot_Legend, JobTrnCrossRef,pollDict,AllSlots,senList=pullTbls(FtBook,TempBook,AssnBook,PollBook)
    #GenerateMasterPollTbl to facilitate making the Slots... require having a table with all employee preferences.
    generateMasterPollTbl(pollDict)
    #Generate Worker Objects, and assign to dictionary keyed by eeID (numeric key, not string keys)
    eeDict=makeEEdict(ftInfoTbl,tempInfoTbl,wkHrs)
    #Generate Schedule Slot objects (all unassigned slots for weekend)
    allSlots=makeSlots(eeDict,AllSlots)
    return Schedule(Acrew,allSlots,eeDict,AssignmentsTbl,senList,pollDict,slot_Legend,pNT=pNT,assnWWF=assnWWF,pVol=pVol,xtraDays=xtraDays,maxI=maxI)


def getEEinfo(FtBook,TempBook):  #Need to make volunteer shift data puller
    """Generate employee objects so as to be able to use their names to read pre generated schedule template."""
    a=getFTinfo(FtBook) #to sqlite
    b=getFTskills(FtBook) #to sqlite
    b=[[int(d[0]),d[1]] for d in b] #Cast EEid to numeric value
    c=getTempinfo(TempBook) #to sqlite
    d=getTempskills(TempBook) #to sqlite
    d=[[int(data[0]),data[1]] for data in d] #Cast EEid to numeric value
    
    addTBL("FTinfo",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=a,addOn=False)
    addTBL("TempInfo",fields=['sen','crew','id','last','first','ytd','totref','totchrg','wtdOT'],data=c,addOn=False)

    return a,c

def addRecs(flNm,shNm,tblNm,data,otptNm='AutoPrimed_Template.xlsx'):
    """Adds data to existing excel table in new rows. Used to flesh out tables in visual template (blank tables)"""
    wb=pyxl.load_workbook(flNm)
    ws=wb[shNm]
    t=ws.tables[tblNm]
    ref=t.ref
    row,col=ws[ref[ref.index(':')+1:]].row,ws[ref[:ref.index(':')]].column
    records=data
    for rec in records:
        for i in range(len(rec)):
            ws.cell(column=col+i,row=row+records.index(rec)+1).value=rec[i]
    newT=pyxl.worksheet.table.Table(displayName=t.displayName,ref=t.ref[:-len(str(row))]+str(row+len(records)))
    style = pyxl.worksheet.table.TableStyleInfo(name="TableStyleLight1",showRowStripes=True)
    newT.tableStyleInfo = style
    # del ws.tables[tblNm]
    # ws.add_table(newT)
    ws.tables[tblNm]=newT
    wb.save(filename=otptNm)


def translate_Visual_Template(flNm,ftRef=None,tRef=None):
    """Takes in an assignment list file, and composes the All_Slots table and Assignment_List table by reading the Visual Template"""
    ftInf,tInf=getEEinfo(ftRef,tRef)
    eeDict=makeEEdict(ftInf,tInf,tp='nm')
    # return eeDict
    wb=pyxl.load_workbook(flNm)
    ws=wb['Visual_Template']
    #===========================================
    #Print out All_Slots table. Simply observe which jobs are present
    data=[]
    jbNms={}
    maxR=5
    for i in range(5,100): #Based on template job names start at row 5
        if ws['A'+str(i)].value!=None: #Typically just 25 jobs, should be extra, so skip blanks
            data.append([1,24,ws['A'+str(i)].value,'','','',1])
            jbNms[i]=ws['A'+str(i)].value #key job name by row id
            if i>maxR: maxR=i #Track row of last job assigned
    addRecs(flNm,'All_Slots','All_Slots',data)
    #===========================================
    #Print out Assignment_List
    data=[]
    got=[]
    #First, identify all slots that are part of a merged cell, so we know to skip over those coordinates when we get to them when iterating over every single one
    for mRng in ws.merged_cells.ranges:
        rw,cl=next(mRng.cells) #Grab the first cell coordinates within merged range as this is where text is stored
        if rw>4: #Do not capture title etc.
            if ws.cell(row=rw,column=cl).value==None:
                pass #if a blank cell was left merged, skip it. Therefore it will be assigned as normal
            else:
                got.extend(mRng.cells) #Gather up merged coords into got tracker
                if ws.cell(row=rw,column=cl).fill==pyxl.styles.PatternFill(fill_type="solid",start_color='FFCC99FF',end_color='FFCC99FF'):
                    tp='F'
                elif ws.cell(row=rw,column=cl).fill==pyxl.styles.PatternFill(fill_type="solid",start_color='FF00B0F0',end_color='FF00B0F0'): 
                    tp='WWF'
                elif 'N/A' in ws.cell(row=rw,column=cl).value:
                    tp='DNS'
                else: tp='V'
                #Active=1, assnType,start slot, end slot, eeid, job
                #slots are -1 because col 1 in excel is job name. Unkn if 2 or 3 cells mrged so use first cell cooridnate already captured, and known that last cell is last entry in got list
                if tp=='DNS':
                    data.append([1,tp,cl-1,got[-1][1]-1,"",jbNms[rw]])
                else:
                    if ' ' in ws.cell(row=rw,column=cl).value: #Pull name to use as dictionary key as anything before appearance of a space 
                        nm=ws.cell(row=rw,column=cl).value[:ws.cell(row=rw,column=cl).value.index(' ')].lower() 
                    else: nm=ws.cell(row=rw,column=cl).value.lower() #No space, use name as appears
                    try:
                        data.append([1,tp,cl-1,got[-1][1]-1,eeDict[nm].eeID,jbNms[rw]])
                    except KeyError:
                        print('The name '+nm+' identified in the visual template could not be associated with EE data from the refusal sheets. Check the formatting')
    #Note that, at this time, the output file already exists with previously added rows in first table. therefore refer to this file as the file to write to, in following command.
    addRecs('AutoPrimed_Template.xlsx','Assignment_List','Assn_List',data)
    # wb.save('wow.xlsx')
    # new=deepcopy(wb)
    # new.save('AutoPrimed_Template.xlsx')
    return 'AutoPrimed_Template.xlsx'