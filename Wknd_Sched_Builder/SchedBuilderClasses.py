from copy import deepcopy
import openpyxl as pyxl
# import functools
import SchedBuilderUtyModule as tls
# from importlib import reload

# def debug(func):
#     """Print the function signature and return value"""
#     @functools.wraps(func)
#     def wrapper_debug(*args, **kwargs):
#         args_repr = [repr(a) for a in args]                      # 1
#         kwargs_repr = [f"{k}={v!r}" for k, v in kwargs.items()]  # 2
#         signature = ", ".join(args_repr + kwargs_repr)           # 3
#         print(f"Calling {func.__name__}({signature})")
#         value = func(*args, **kwargs)
#         print(f"{func.__name__!r} returned {value!r}")           # 4
#         return value
#     return wrapper_debug


class Slot():
    """A single 4 hour time slot for a single job, to be filled by 1 person"""
    def __init__(self,seqID,dispNm,trnNm=None):
        self.trnNm=trnNm #to be used when filtering out staff for training
        self.dispNm=dispNm #to be used for printouts
        self.seqID=seqID
        #self.datetime=   #Determined based on seq. Used in printout of assignments
        self.assignee=None #To store eeid of assignee for printout
        self.assnType=None #e.g. Forced/WWF
        self.slotInShift=0 #1 if first slot in someones shift. 2 if second, etc.
        self.totSlotsInShift=0 # 1 if 4 hours shift, 2 if 8 hour shift, 3 if 12 hour shift
        self.eligVol=0 #This will be used to track which slot is the most constrained...it tracks the count of how many people eligible volunteers this slot has going for it
        self.disallowed=[] #EEID's that were specified as not allowed to be assigned to this slot in the Assn List

    def key(self):
        return str(self.seqID)+'_'+self.dispNm
    
    def assn(self,sch,assnType=None,slAssignee=None,fromList=False):
        """Assign a slot to someone, and perform associated variable tracking etc. Returns a bool indicating if a forcing rule was broken or not"""
        if (slAssignee is not None) and assnType=='DNS': #Case that this is specifying *not* to assign someone. In every other case it is a matter of actually assigning someone 
            self.disallowed.append(slAssignee)
        else:
            self.assnType=assnType
            self.assignee=slAssignee #eeid
            if assnType in ['WWF','F','V']:
                pass
                # del sch.slots[self.key()] #Remove this slot from the 'openslots' collection if someone was actually assigned
                # #del sch.slots[self.key()]
                # del sch.fslots[self.key()]
            elif assnType=='nV':
                pass
                #del sch.slots[self.key()]
            if slAssignee is not None: #Case of specific assignment, only not follwoed through when its no ee and DNS
                sch.ee[slAssignee].assnBookKeeping(self,sch) #add this slot to the ee's assigned slot dictionary & other tasks
        #Logging for printout after
        if assnType!=None and slAssignee!=None:
            logTxt=''
            if fromList==True:
                logTxt+= 'Per Assn List: '
            if assnType=='DNS':
                logTxt+='Removed slot '+self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+') from scheduling'
                if slAssignee is not None: logTxt+= ' for ee '+ sch.ee[slAssignee].firstNm[0]+'. '+sch.ee[slAssignee].lastNm
            elif assnType=='WWF': logTxt+="WWF Assignment: "+sch.ee[slAssignee].firstNm[0]+'. '+sch.ee[slAssignee].lastNm+' to ' +self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+')'
            elif assnType=='F': logTxt+="   FORCED Assignment: "+ sch.ee[slAssignee].firstNm[0]+'. '+sch.ee[slAssignee].lastNm+ ' to ' +self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+')'
            elif assnType=='V': logTxt+="   Voluntary Assignment: "+ sch.ee[slAssignee].firstNm[0]+'. '+sch.ee[slAssignee].lastNm+' to ' +self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+')'
            elif assnType=='N':logTxt+="   No voluntary or forced assignment could be made to "+self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+')'
            elif assnType=='nV':logTxt+="   No voluntary assignment could be made to "+self.dispNm+' '+ sch.slLeg[self.seqID-1][2]+' ('+sch.slLeg[self.seqID-1][1]+')'
            sch.assnLog.append(logTxt)
        if slAssignee!=None and assnType in ['V','F']: #Check if assignment breaks forcing rules
            if sch.ee[slAssignee].frcOK(sch)!=True: #Case that the assignment broke a forcing rule... that slot needs be earlier priority.
                return False
        return True

class ee():
    """A staff persons data as related to weekend scheduling"""
    def __init__(self,snty,crew,id,Last,First,refHrs,wkHrs,wkndHrs=0,skills=[]):
        self.seniority=int(snty)
        if refHrs==None:
            self.refHrs=0
        else:
            self.refHrs=float(refHrs)
        self.wkdyHrs=float(wkHrs)
        self.wkndHrs=float(wkndHrs)
        self.frcHrs=0
        self.lastNm=Last
        self.firstNm=First
        self.eeID=int(id)
        self.crew=crew
        self.assignments=[]#To be appended with slots as they are assigned, keyed as they are in the slot dictionary
        self.skills=skills 
    
    def dispNm(self,slt=None):
        if slt is None:
            if int(self.seniority)>50000: return self.firstNm[0]+'.'+self.lastNm[0]+self.lastNm[1:].lower()+'(T)' #Case of Temps
            else: return self.firstNm[0]+'.'+self.lastNm[0]+self.lastNm[1:].lower()
        else: #Can make functionality to pass in 'MESR' to display name based on some slot criteria?
            pass

    def frcOK(self,sch):
        """Returns if the present assignments are permissible with rules around forcing limitations"""
        #Used after making assignments to check if need make a slot priority or not
        asn=sorted(self.assignments,key=lambda k:int(k[:k.index('_')])) #Order assignment keys by their slot ID's
        h=self.wkdyHrs #initialize tally
        for k in asn: #For each slot, considering the hours worked upon completion of that workslot
            h+=4
            if h>48 and sch.slots[k].assnType=='F': return False #Forcing rules broken, if condition is met then forcing past 48 hours in week
        return True #if the above condition not met.. all good    

    def assnBookKeeping(self,sl,sch):
        """Carried out when assigned to slot, adjusts tally of eligible volunteers to other slots accordingly"""
        self.assignments.append(sl.key())
        if sl.assnType=='F': self.frcHrs+=4 #Track forced hours
        self.wkndHrs+=4
        #   Return keys for slots that the person was counted as an eligible volunteer for
        kys=[k for k in sch.slots if self.eeID in sch.slots[k].eligVol]
        for k in kys:
            if self.slOK(sch,sch.slots[k],pt=False,poll=tls.viewTBL('allPollData',filterOn=[('eeid',self.eeID)])[0]) is not True: sch.slots[k].eligVol.pop(sch.slots[k].eligVol.index(self.eeID)) 
            #pop the eeId out of eligVol list if the given slot is no longer ok to assign
        #The slOK function does not capture if making a voluntary assignment earlier in the weekend invaldiates a forced assignment later in the weekend.
        #That is captured in a separate function 'frcOK'

    
    def totShiftHrs(self,sl,toFlw=False,styling=False):
        """Given a slot, assuming it is assigned, what is the total shift length of the shift in which that slot is a constituent. If toFlw=True then return # slots to follow present slot in same shift"""
        sLen=1 #start off shift length at one because the slot being passed in is always minimum
        assnSeqIDs=[int(k[:k.index('_')]) for k in self.assignments] #Pull out the seqID's from the key strings for each slot an ee is already assigned
        if len(assnSeqIDs)==0:
            if toFlw==False:
                return 4 #Case that no slots assigned so far, so if sl assigned then its alone
            else: return 0 #single shift, no following
        else:
            if toFlw==False:
                anch=sl.seqID
                assnSeqIDs.append(anch)
                assnSeqIDs.sort() #sorts integers lowest to highest
                if styling==True: #Coutning method is simpler when just counting shift length for schedule colouring, need not account for slot passed in being potential 4slot consectuvie work slot
                    for offset in [-2,-1,1,2]: #Knowing that shifts will never be assigned 4 in a row, and there won't be a 4 hour gap between shifts, just count neighbouring 4 slots (2 on each side)
                        if anch+offset in assnSeqIDs: 
                            sLen+=1
                    return sLen*4 #Return num hours
                else: #Originally was using the same method as styling, with limits as -3,+3, but made a bug.. even if the two neighbouring slots on 1 side were unassigned (8hr gap), it would still count the 3rd if it was assigned, even though it shouldnt be counted as not a contiguous shift
                    assnSeqIDs=list(set(assnSeqIDs)) #converting to set eliminates duplicates ( in event slot being passed in was already assigned.. it was appended again)
                    def ranges(nums):
                        """Returns a list of (open,close) intervals a list spans"""
                        nums = sorted(set(nums))
                        gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s+1 < e]
                        edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
                        return list(zip(edges, edges))
                    rns=ranges(assnSeqIDs) #dont have to worry that following line failes to index list comp @[0] because the slot in question is always passed in, even if no other assignments, the slot in question will be returned as 1 slot 4 hours
                    myRn=[x for x in rns if x[0]<=anch and x[1]>=anch][0] #Pull out a tuple contianing lower and upper bound of sequence of seqID's of which the slot in question is a part
                    return (myRn[1]-myRn[0]+1)*4 #e.g. range is seqIDs (6,7), thats two slots. 7-6+1=2
            else: #Count # of slots that follow this one consecutively
                i=0 #Count of slots to follow on same shift
                for offset in [1,2]: #Never more than 3 in a row so need only check the 2 following slot seqID's for assignment
                    if anch+offset in assnSeqIDs: 
                        i+=1
                return i
    
    def assnConflict(self,sl):
        """Returns true if someone is already assigned to a slot with same seqID as potential assignment, false if no conflict"""
        assns=[int(k[:k.index('_')]) for k in self.assignments] #Pull out the seqID's form the key strings for each slot an ee is already assigned
        if len(assns)==0:return False #if no other assignments.. no conflict
        elif sl.seqID not in assns: return False #if no other assns with same seqID.. no conflict
        elif sl.seqID in assns: return True #if other assignment already amde with same seqID... true, conflict present

    
    def gapOK(self,sl,sch,tp='V'):
        """Returns true if the slot, when assigned, doesn't break the rule for minimum gap between shifts. 12 hours for forcing, 8 for vol"""
        assns=[int(k[:k.index('_')]) for k in self.assignments] #Pull out the seqID's form the key strings for each slot an ee is already assigned
        #Just need to check the nearest neighbours aren't with a gap of 1 empty slot
        deltaNextShifts=[v for v in [sId-sl.seqID for sId in assns] if v>0]  # Define this and the next before defining next/prevNeighDist to manage case of 0 this being equal to [], which min() can't accept
        deltaPrevShifts=[v for v in [sl.seqID-sId  for sId in assns] if v>0] ##Basically, subtract various Id's from sl in question. Remove vlaues<0 bc those slots come later (next neighb). Then take the minimum value, which is the diff in seqID between sl being compared, and nearest neighbour on left when sorted chronologically 
        if len(deltaNextShifts)==0:
            deltaNextShifts=[0] #Fill with bogus # to give 'ok' result to function
        if len(deltaPrevShifts)==0:
            deltaPrevShifts=[0]
        nextNeighbDist=min(deltaNextShifts)
        prevNeighbDist=min(deltaPrevShifts) 
        #Utility functions:
        def okForLastWkShift():
            if self.crew==sch.Bcrew and sl.seqID-6*(sch.friOT==False)<3+1*(tp=='F'):
                return False #Case of B shift worker being assigned within 8 (or 12 if forced) hrs of last weeks final afternoon shift, no go
            elif self.crew==sch.Acrew and tp=='F' and sl.seqID-6*(sch.friOT==False)<2:
                return False #Case of A shift worker being forced onto first slot of the weekend
            else: return True
        def okForNextWkShift():
            if self.crew=='Rock' and tp=='V' and sl.seqID+6*(sch.monOT==False)==23:
                return False #Case of night shift person being assigned 3p-7p afternoon before weekday night shift
            elif self.crew=='Rock' and tp=='F' and sl.seqID+6*(sch.monOT==False)>21: 
                return False #Case of night shift forcing.. can't be forced such that <12 hours before next shift
            elif self.crew==sch.Acrew and tp=='F' and sl.seqID+6*(sch.monOT==False)==24:
                return False #Case of day shift ee (next week) being forced 7p-11p the night before
            else: return True
        #=====
        if tp=='V' and (nextNeighbDist==2 or prevNeighbDist==2):
            return False#Distance of one means consecutive slots (shiftlength already checked), distance greater than 2 is gap of 8 hours or more
        elif tp=='F' and ((nextNeighbDist==2 or prevNeighbDist==2) or prevNeighbDist==3):return False #Forcing requires gap of 12h or more going into it, 8 hours after it
        elif okForLastWkShift()==True and okForNextWkShift()==True: return True #Reaching this elif means that the other conditions aren't true, so lastly just have to check the gap with weekday shifts ok
        else: return False #Some condition not met       

    def slOK(self,sch,sl,poll=0,tp='V',pt=True): #pt is 'print'... 
        """Returns True if the slot being tested is ok to be assigned, false if not"""
        #Test all conditions (trained, wk hrs, consec shift, time between shifts, before making a branch to test willingness or not based on assignment type forced/voluntary)
        if (sl.dispNm in self.skills): #the person is trained and hasn't been specified in assignment log *not* to be assigned here
            if (self.eeID not in sl.disallowed): #this ee/slot pairing isn't ruled out in disallowment list
                if self.wkndHrs+self.wkdyHrs<60: #total week hours ok!
                    if self.assnConflict(sl)==False: #No existing assignment at same time!
                        if self.totShiftHrs(sl)<=12: #This slot wouldn't have a given shift exceed 12 hours
                            if self.gapOK(sl,sch,tp=tp): #This slot being assigned doesn't break a shift gap rule
                                if self.crew in ['wwf','bud','blue','rock','silver','gold','student']:
                                    if tp=='V':#voluntary: check willigness
                                        if (poll[3+sl.seqID] !="") and (poll[3+sl.seqID] is not None) and (poll[3+sl.seqID]!='n'): #Person is willing!
                                            return True
                                        else: 
                                            if sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Did not volunteer for this shift')
                                            return False
                                    else: 
                                        if self.wkndHrs+self.wkdyHrs<48: return True #Forced
                                        elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Cannot force past 48 hours worked in week')
                                elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Crew is not in list wwf,bud,blue,rock,silver,gold,student')
                            elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Insufficient gap time between this shift and another')
                        elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Total consecutive hours would exceed 12')
                    elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Is already assigned for slot in same time period')
                elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Total hours in week exceeds 60')
            elif sch.sF==False and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Disallowed in Assignment List')
        elif sch.sF==False and sch.pNT==True and pt==True: sch.assnLog.append('   Fail to assign '+self.firstNm[0]+'. '+self.lastNm+' to ' +sl.dispNm+ ' '+sch.slLeg[sl.seqID-1][1]+' '+sch.slLeg[sl.seqID-1][2]+' || Not Trained')
        return False

class Schedule():
    def __init__(self,Acrew,slots,ee,preAssn,senList,polling,slLeg,sF=False,pNT=False):
        # self.ftInfoTbl=ftInfoTbl
        self.pNT=pNT #if True prints '...Not Trained'' statement as applicable when testing if a given slot is ok for someone
        self.sF=sF #suppressFails.. if false, prints out 'failed to schedule' statements
        self.Acrew=Acrew
        if Acrew=='Bud':self.Bcrew='Blue'
        else:self.Bcrew='Bud'
        self.slots= slots  #A collection of Slot objects that compose this schedule
        # self.slots= deepcopy(slots) #Referenced in forcing phase 1
        # self.slots=deepcopy(slots) #Referenced in voluntary assignment phase
        # self.fslots=deepcopy(slots) #Referenced in forcing phase 2
        self.ee=ee #A dictionary containing ee info
        self.preAssn=preAssn #A list of lists containing the predefind assignment info
        self.senList=senList
        self.polling=polling
        self.assnLog=[] #To be appended when assignments made, for read out with final product
        self.slLeg=slLeg #Slot Legend. Used for easy refernece of slot times after.
        seqIDs=[int(k[:k.index('_')]) for k in self.slots]
        #look at slots to see if fri/mon OT to be assigned... used in 'gapOK' method for ee
        if min(seqIDs)>6: self.friOT=False
        else: self.friOT=True
        if max(seqIDs)<19: self.monOT=False
        else: self.monOT=True 
        self.rev=0
        self.noVol=[] #A list to contain keys of slots with no eligible volunteers.
    
    def evalAssnList(self):
        """Enter all predefined assignments into the schedule"""
        #First, iterate through the assignment list. Each record will generate one or more records for the 'slot change log'
        #The slot change log will have one record for each slot for which an assignment (or other specified status change) should be made
        #So assignment log records that indicate a span of multiple slots will generate multiple records in the change log.
        slChLg=[] #Initialize slot change log. It will be a list of lists where each sublist has the necessary info within to pass to the slot assignment function
        #===
        def getKeys(seqId_1,seqId_2,jobNm=None):
            """Returns the keys for all slots that a given assnLog record applies to, job specified or not"""
            myKeys=[]
            if jobNm==None:
                for seqNo in range(seqId_1,seqId_2+1):#Apply to all slots in given range.. +1 due to range fn not being inclusive
                    moreKeys=[k for k in self.slots.keys() if k[:len(str(seqNo))+1]==str(seqNo)+'_'] #Pull dict keys for Slots where it is a slot with matching seqNo, regardless of job name
                    myKeys.extend(moreKeys)
            else: #job is defined
                for seqNo in range(seqId_1,seqId_2+1):
                    myKeys.append(str(seqNo)+'_'+jobNm)
            return myKeys
        #===
        #Here we get down to business - Reading the assignment log, generating records in the slotChangeLog, and then evaluating those changes
        for myAssn in self.preAssn:
            if myAssn[0]==1: #Only evaluate those with '1' in 'Active' (first) column
                assnTp=myAssn[1]
                if (myAssn[5]=="" or myAssn[5]==None): jb=None #grab job name
                else: jb=myAssn[5]
                if (myAssn[4]=="" or myAssn[4]==None): asgne=None #grab assignee
                else: asgne=myAssn[4]
                keys=getKeys(myAssn[2],myAssn[3],jb) #pull all the keys for slots this particular assn list item applies to
                for k in keys:
                    slChLg.append([k,self,assnTp,asgne]) #Add record(s) to the slot change long, one for each record.
        #Now that the slChLg is made, carry out the function that reads it record by record and goes and modifies the slots
        def evalLogRec(rec):
            """Carry out the 'assn' method on the associated slot with relevant data from Assn log"""
            self.slots[rec[0]].assn(rec[1],rec[2],rec[3],fromList=True)
        for rec in slChLg:
            evalLogRec(rec)
    
    def proofEligVol(self):
        """This clears the eligVol lists for all slots of the eeID's where the person isn't eligible"""
        #Necessary because had to make slots before making schedule, so wasn't able to actually test if slOK before assigning ee to slot when initializing everything
        for k in self.slots:
            s=self.slots[k]
            for e in s.eligVol:
                if self.ee[e].slOK(self,s,poll=tls.viewTBL('allPollData',filterOn=[('eeid',e)])[0],pt=False) is not True:
                    s.eligVol.pop(s.eligVol.index(e))

    def nextSlots(self,force=0):
        """Returns the next most constrained unassigned slot object. If 'forcing'=True then returns list of slots with 0 eligible assignes, ordered by seqID"""
        if force==0: #Proceed with selecting most constrained slot with >=1 potential assignees
            kyNonZero=[s for s in self.slots if (len(self.slots[s].eligVol)>0) and self.slots[s].assnType not in ['WWF','F','V','nV','DNS','N'] ] #Get keys for slots with >0 eligVol
            if len(kyNonZero)==0: return None #if none left to assign, just return None
            eligCnts=[len(self.slots[k].eligVol) for k in kyNonZero]
            # print(list(zip([self.slots[s].key() for s in kyNonZero],eligCnts)))
            if eligCnts.count(min(eligCnts))>1: #Case that there are slots tied for most constrained
                slts=[self.slots[s] for s in kyNonZero if len(self.slots[s].eligVol)==min(eligCnts)] #Retrieve the tied slot objects
                totSkills=[sum([len(self.ee[eId].skills) for eId in s.eligVol ]) for s in slts ] #For each slot, make a list of integers, whee each integer is the number of jobs an ee eligible for that slot is trained on. Sum those lists, and the slot with the highest number is selected, since that is correlated with the eligible assignees for that slot having the most ability to cover other slots.
                if totSkills.count(min(totSkills))>1: #Case that 2 slots are tied for minimum of total # of training records for eligible operators
                    #Go with the one for which there is an operator with least spots trained... assuming that the operator who is most constrained training wise gets it.. while this is an assumption without great basis, there is at least the point that someone with less training will likely have less refusal hours in the year.. so it may turn to work out ok
                    trainRecForLeastTrainedEE=[min([len(self.ee[eId].skills) for eId in s.eligVol ]) for s in slts ] #Same formula as totSkills except min instead of sum
                    pickSl=slts[trainRecForLeastTrainedEE.index(min(trainRecForLeastTrainedEE))]
                    self.assnLog.append('Slot '+pickSl.key()+' (assnType: '+str(pickSl.assnType)+') chosen as most constrained, was tied for totSkills, chose first')
                    return pickSl #Here only 1 return statement because if its a tie we'll just take the first one, which the index function here will give.
                else: 
                    pickSl=self.slots[ kyNonZero[totSkills.index(max(totSkills))]]
                    self.assnLog.append(pickSl.key()+' (assnType: '+str(pickSl.assnType)+') chosen as most constrained, had most TotSkills')
                    return pickSl #Case of one slot having more totSkills than another. Call slots keys, then index that by the totSKills count tog et the index of the slot we want, and retrieve that from slots, 
            else:
                pickSl=self.slots[kyNonZero[eligCnts.index(min(eligCnts))]]
                self.assnLog.append(pickSl.key()+' (assnType: '+str(pickSl.assnType)+') chosen as most constrained, had least ('+str(len(pickSl.eligVol))+') eligible volunteers')
                return pickSl #Retrieve most constrained if not tied with any other.
        elif force==1:#Forcing for the first time. Return list of slots to force into in chronological order
            return sorted([self.slots[s] for s in self.slots if len(self.slots[s].eligVol)==0 and self.slots[s].assnType == None],key=lambda x: x.seqID)
        elif force==2: #Forcing for teh 2nd time. Return all slots. the 'eligibility tracking' isn't perfect so can't filter by it because when people were assigned, would be a pain to make logic to properly remove their 'eligVol' status from the slots which they were no longer eligible for for reasons like max shift length etc. Only removed it for slots happenign at same time
            return sorted([self.slots[s] for s in self.slots if self.slots[s].assnType == 'nV' or self.slots[s].assnType == None],key=lambda x: x.seqID)

    def pickAssignee(self,sl,tp='V',pt=True):
        """Returns an eeid and the assignment type, either voluntary or forced, or 'N" for None/No staff, for the passed slot"""
        if tp=='V':
            def tblSeq(sl,Acrew,Bcrew):
                """Returns keys for retrieving poll data tables in sequence of priority assignment. With respect to shift, assignment priority goes in CABCA sequence, first FT then Temps"""
                if sl.seqID in [1,2,7,8,13,14,19,20]: homeShift='C'
                elif sl.seqID in [3,4,9,10,15,16,21,22]: homeShift='A'
                else: homeShift='B'
                keys=[]
                seqStr='CABCA'
                cD={'C':'Rock','A':Acrew,'B':Bcrew} #Crew Dict
                for eeTp in ['FT','Temp']: #Go through all FT's before temps
                    for i in range(3):#code below uses variable homeShift in conjunction with stepping through seqStr to pull out crews in order of priority selection of OT
                        keys.append('tbl_'+cD[seqStr[seqStr.index(homeShift)+i]]+eeTp)
                return keys
            #===
            ks=tblSeq(sl,self.Acrew,self.Bcrew)
            #Relying on the fact that the tables in the excel sheet were already sequenced in order of refusal hours...
            for k in ks:#Iterate through the tables in provided sequence to pull from crews in sequence of priority pick
                for rec in self.polling[k]: #Iterate through rows in table to pull eeID's in sequence of hours
                    if rec[0] is not None: #Error proof on having an empty polling table for a particular crew
                        if self.ee[rec[0]].slOK(self,sl,poll=tls.viewTBL('allPollData',filterOn=[('eeid',rec[0])])[0],pt=pt):
                            return rec[0],'V' #Person has been found 
                        else: pass
            return None,'nV' #No voluntary assignee found                                
        elif tp=='F':
            if self.pickAssignee(sl,pt=False)[0]!=None: return self.pickAssignee(sl,pt=False) #Check that someone is willing. This branch can be reached in phase 1 forcing when recursing because one force created no takers in previous slot, and now testing to force a slot of different time code. Possible that someone may be willing but shift gap doesnt allow forcing, so need this statement to test on voluntary though command is called from forcing phase
            for i in range(len(self.senList)-1,-1,-1): #Work way down seniority list
                lowManID=int(self.senList[i][2])
                if self.ee[lowManID].slOK(self,sl,tp='F'): return lowManID,'F'
            return None,'N' #No one to force

    # @debug
    def fillOutSched_v2(self,noVol=None,iter=0):
        """Fills out schedule in a recursive way... see blog... when filling with voluntary folks, if a slot with no takers is encountered and has never been identified as such before, then restart from a fresh sched seed where that slot is forced before any voluntary assignments happen"""
        iter+=1
        WIPschd=deepcopy(self) #WIPschd will have assignments made to it. 'Self' is kept with only the AssnList stuff coming into this point so that across multiple iterations where the NoVol list is potentially expanded, it serves as the blank slate 
        if iter==1: #First iteration. Initialize noEligVol list. Do not define it this way in future iterations, because you will lose the slots that were discovered that needed to be added!
            WIPschd.noVol=[k for k in WIPschd.slots if len(WIPschd.slots[k].eligVol)==0 and WIPschd.slots[k].assnType not in ['DNS','WWF','V','F']]
        else: #2nd and further iterations.. coming here because more slots were found to force and would've passed along in function call... so retrieve them here
            WIPschd.noVol=noVol
        #===Logging for printout
        WIPschd.assnLog.append('Iteration: '+str(iter)+' ||  Starting Schedule With Identified Priority Slots, Forcing As Necessary:')
        mynoVol=''
        for x in sorted(WIPschd.noVol,key=lambda k:int(k[:k.index('_')])): mynoVol+=' '+x
        WIPschd.assnLog.append('Prelim Priority Assignment Sequence: '+mynoVol)
        #=====
        #Initial phase.. Volunteer assignments or force as necessary for priority slots identified via no volunteers, or became no-vol on previous iterations
        #       Slots found to lose all eligVol on prev iterations added to this bunch
        #       Also need to check if making a forcing creates the need to make another forcing, and perform recursion in that case as well!
        #       Because the forcing list has slots which were determined via prior recurses in the scheduling that identified a dead end requiring forcing, can't use the same method for tracking new slots needing forcings as used for voluntary assignments
        for k in sorted(WIPschd.noVol,key=lambda k: int(k[:k.index('_')])): #Iterate through the keys in chronological order          
            s=WIPschd.slots[k]
            WIPschd.assnLog.append('Looking to Force to '+s.dispNm+' '+ WIPschd.slLeg[s.seqID-1][1]+' '+ WIPschd.slLeg[s.seqID-1][2])
            eId,tp=WIPschd.pickAssignee(s,tp='F')
            r=s.assn(WIPschd,assnType=tp,slAssignee=eId)
            if r==False and s.key() not in WIPschd.noVol:
                WIPschd.noVol.append(s.key())
                return self.fillOutSched_v2(WIPschd.noVol,iter) #WIPschd,'P-Frce Brk' #<-Alt return for debugging
            newK=set([k for k in WIPschd.slots if len(WIPschd.slots[k].eligVol)==0 and WIPschd.slots[k].assnType not in ['WWF','F','V','nV','DNS','N']]) #After assignment, see if anything now needing forcing that hasn't been seen before
            if len(newK-set(WIPschd.noVol))>0: #Case that a forced assignment made someone ineligible for a slot they were marked as a volunteer in
                pullK=newK-set(WIPschd.noVol) #Get the keys for slots that are now without volunteers(could be more than one so use set subtraction)
                WIPschd.noVol.extend(pullK)
                WIPschd.assnLog.append('The last assignment resulted in 1 or more other slots having no more eligible volunteers. Those slots are added to the list of slots to force at the start, and a new schedule will be made with updated list of slots to Force')
                self.assnLog.extend(WIPschd.assnLog)
                return self.fillOutSched_v2(WIPschd.noVol,iter) # WIPschd,'P-Bump'  #<-Alt return for debugging

        #Second phase - Follow most constrained slot to amke assignments
        #      If a slot gets reduced to no eligible volunteers, track it to the priority list and restart
        OldToAssn=[k for k in WIPschd.slots if (len(WIPschd.slots[k].eligVol)>0) and WIPschd.slots[k].assnType not in ['WWF','F','V','nV','DNS','N'] ]
        NumSl=len(OldToAssn)
        #===========Printouts for Reporting
        appn='Second Phase Assignments (Sequence by Most Constrained): '+str(NumSl)+' slots  ||  '
        for x in OldToAssn: appn+=x+' - '
        WIPschd.assnLog.append(appn)
        for i in range(NumSl): #Iterate across all slots identified at the start
            NewToAssn=[k for k in WIPschd.slots if (len(WIPschd.slots[k].eligVol)>0) and WIPschd.slots[k].assnType not in ['WWF','F','V','nV','DNS','N'] ]
            #===========Printouts for troubleshooting
            # appn='Iter: '+str(i)+' | OldToAssn: '
            # for x in OldToAssn: appn+=x
            # appn+='  | NewToAssn: '
            # for x in NewToAssn: appn+=x
            # WIPschd.assnLog.append(appn)
            #============
            slLost=set(OldToAssn)-set(NewToAssn)
            if len(slLost)<2:
                #Case that this is iteration one (old-new=0) OR this is the expected most common case of the last assignment having been made means that NewToAssn is only missing that last slot, as compared to the Old. If >=2, this means that the last assignment made has resulted in no eligible volunteers for a slot, so there are 2 or more slots missing from NewToAssn list
                OldToAssn=NewToAssn #consider the new as old for next iteration
                curS=WIPschd.nextSlots() #Pick most constrained of all avail slots
                lastK=curS.key() #In case needed to remove this slot from set of keys to add to WIPschd.noVol in next iter if multiple slots made to require forcing
                if curS is not None:
                    WIPschd.assnLog.append('Looking to voluntarily assign to '+curS.dispNm+' '+ WIPschd.slLeg[curS.seqID-1][1]+' '+ WIPschd.slLeg[curS.seqID-1][2])
                    eId,tp=WIPschd.pickAssignee(curS)
                    r=WIPschd.slots[curS.key()].assn(WIPschd,assnType=tp,slAssignee=eId) #Note the assign method acts on original, not deepcopy, retrieved via key
                    if r==False and curS.key() not in WIPschd.noVol:
                        WIPschd.noVol.append(curS.key())
                        return self.fillOutSched_v2(WIPschd.noVol,iter) # WIPschd,'V-F Rule'   #<-Alt return for debugging
                if i==NumSl-1:return WIPschd#,'WIN' #Once all voluntary assignments made, schedule done? May need to force more, or go through force function to designate 'no staff' slots
            else: #Case that assignment being made resulted in more slots being left with no EligVol.. add those slots to noVol list and re start the function
                WIPschd.assnLog.append('The last assignment resulted in 1 or more other slots having no more eligible volunteers. Those slots are added to the list of slots to force at the start, and a new schedule will be made with updated list of slots to Force')
                WIPschd.noVol.extend([k for k in slLost if k!=lastK]) #lastK got assigned, so remove it from slLost to get all slots that are identified as needing to be forced due to person last assigned not being available for other slots...
                return self.fillOutSched_v2(WIPschd.noVol,iter)# WIPschd,'V-Bump' #<-Alt return for debugging
        #Third phase
        # 3. Final Forced Filling
        # WIPschd.assnLog.append('Final Forcing Phase... Forcing to slots with no more eligible volunteers after having assigned voluntary OT')
        # sls=WIPschd.nextSlots(force=2)
        # for s in sls:
        #     WIPschd.assnLog.append('Forcing Phase 2 on '+s.dispNm+' '+ WIPschd.slLeg[s.seqID-1][1]+' '+ WIPschd.slLeg[s.seqID-1][2])
        #     eId,tp=WIPschd.pickAssignee(s,tp='F')
        #     s.assn(WIPschd,assnType=tp,slAssignee=eId)

    def fillOutSched(self):
        """Having made the predetermined assignments, fill in the voids in the schedule"""
        #Algorithm is basically:
        # 1. Force staff for slots with no eligible assignees
        # 2. Iterate through unassigned slots in sequence of which is most constrained
        #    Assign staff in order of who gets priority pick at the slot
        # 3. Force for slots that had no eligible after giivng the eligble their voluntary choices
        #    If no forcing availability, label as such and move on
        #End when no more unassigned slots left

        #Proceed with carrying out the algorithm:
        # 1. Initial Forcing
        sls=self.nextSlots(force=1)
        self.assnLog.append('Initial Forcing phase... Forcing to slots with no eligible volunteers')
        for s in sls:
            self.assnLog.append('Looking to Force to '+s.dispNm+' '+ self.slLeg[s.seqID-1][1]+' '+ self.slLeg[s.seqID-1][2])
            eId,tp=self.pickAssignee(s,tp='F')
            s.assn(self,assnType=tp,slAssignee=eId)
        # 2. Voluntary Filling
        self.assnLog.append('Voluntary Assignment Phase... Assigning slots in sequence of most to least constrained by number of eligible volunteers')
        sls=deepcopy(self.slots)
        toAssn=[s for s in self.slots if (len(self.slots[s].eligVol)>0) and self.slots[s].assnType not in ['WWF','F','V','nV','DNS','N'] ]
        #toAssn is the same as the first action in nextSlots().. needed here to iterate the right number of times
        for i in range(len(toAssn)): #Iterate across all slots,
            s=self.nextSlots() #Pick most constrained of all avail slots
            if s is not None:
                self.assnLog.append('Looking to voluntarily assign to '+s.dispNm+' '+ self.slLeg[s.seqID-1][1]+' '+ self.slLeg[s.seqID-1][2])
                # sls.pop(s.key()) #Remove that one from set for next iteration
                eId,tp=self.pickAssignee(s)
                self.slots[s.key()].assn(self,assnType=tp,slAssignee=eId) #Note the assign method acts on original, not deepcopy, retrieved via key        
                # print(str(i))
                # print(self.slots['8_Labeler'].assnType)
        # 3. Final Forced Filling
        self.assnLog.append('Final Forcing Phase... Forcing to slots with no more eligible volunteers after having assigned voluntary OT')
        sls=self.nextSlots(force=2)
        # print('force slots:')
        # print([x.key() for x in sls])
        for s in sls:
            self.assnLog.append('Forcing Phase 2 on '+s.dispNm+' '+ self.slLeg[s.seqID-1][1]+' '+ self.slLeg[s.seqID-1][2])
            eId,tp=self.pickAssignee(s,tp='F')
            s.assn(self,assnType=tp,slAssignee=eId)


    def printToExcel(self):
        """Print all slot assignments to an excel file for human-readable schedule interpretation"""
        #Define Cell styling function
        def styleCell(cl,clType,s=None,horizMergeLength=0):
                for i in range(horizMergeLength+1):
                    cl=cl.offset(0,i)
                    if clType=='hours':
                        cl.font=pyxl.styles.Font(bold=True,size=16,color="00FFFFFF")
                        cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='00FF0000',end_color='00FF0000')
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thick'),
                        right=pyxl.styles.Side(border_style='thick'),
                        top=pyxl.styles.Side(border_style='thick'),
                        bottom=pyxl.styles.Side(border_style='thick'))
                    elif clType=='shift':
                        cl.font=pyxl.styles.Font(bold=True,size=16)
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thick'),
                        right=pyxl.styles.Side(border_style='thick'),
                        top=pyxl.styles.Side(border_style='thick'),
                        bottom=pyxl.styles.Side(border_style='thick'))
                    elif clType=='day':
                        cl.font=pyxl.styles.Font(bold=True,size=28)
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thick'),
                        right=pyxl.styles.Side(border_style='thick'),
                        top=pyxl.styles.Side(border_style='thick'),
                        bottom=pyxl.styles.Side(border_style='thick'))
                    elif clType=='DNS':
                        cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='B2B2B2',end_color='B2B2B2')
                        cl.font=pyxl.styles.Font(bold=True,size=14)
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                        right=pyxl.styles.Side(border_style='thin'),
                        top=pyxl.styles.Side(border_style='thin'),
                        bottom=pyxl.styles.Side(border_style='thin'))
                    elif clType=='WWF':
                        cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='00B0F0',end_color='00B0F0')
                        cl.font=pyxl.styles.Font(bold=True,size=14,color="00FFFFFF")
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                        right=pyxl.styles.Side(border_style='thin'),
                        top=pyxl.styles.Side(border_style='thin'),
                        bottom=pyxl.styles.Side(border_style='thin'))
                    elif clType=='F':
                        cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='CC99FF',end_color='CC99FF')
                        cl.font=pyxl.styles.Font(bold=True,size=14)
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                        right=pyxl.styles.Side(border_style='thin'),
                        top=pyxl.styles.Side(border_style='thin'),
                        bottom=pyxl.styles.Side(border_style='thin'))
                    elif clType=='N':
                        cl.font=pyxl.styles.Font(bold=True,size=16,color="00FFFFFF")
                        cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='00FF0000',end_color='00FF0000')
                        cl.alignment=pyxl.styles.Alignment(horizontal='center')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thick'),
                        right=pyxl.styles.Side(border_style='thick'),
                        top=pyxl.styles.Side(border_style='thick'),
                        bottom=pyxl.styles.Side(border_style='thick'))
                    elif clType=='jbNm':
                        cl.font=pyxl.styles.Font(bold=True,size=14)
                        cl.alignment=pyxl.styles.Alignment(horizontal='left')
                        cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                        right=pyxl.styles.Side(border_style='thin'),
                        top=pyxl.styles.Side(border_style='thin'),
                        bottom=pyxl.styles.Side(border_style='thin'))
                    elif clType=='V': #Note that if slot was forced, it will be purple per above if branch, regardless of shift length
                        if self.ee[s.assignee].totShiftHrs(s,styling=True)==4: #Colour 4 hour shifts yellow
                            cl.font=pyxl.styles.Font(bold=True,size=14)
                            cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='00FFC000',end_color='00FFC000')
                            cl.alignment=pyxl.styles.Alignment(horizontal='center')
                            cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                            right=pyxl.styles.Side(border_style='thin'),
                            top=pyxl.styles.Side(border_style='thin'),
                            bottom=pyxl.styles.Side(border_style='thin'))
                        elif self.ee[s.assignee].totShiftHrs(s,styling=True)==8: #Leave 8 hrs shift white
                            cl.font=pyxl.styles.Font(bold=False,size=14)
                            cl.alignment=pyxl.styles.Alignment(horizontal='center')
                            cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='00FFFFFF',end_color='00FFFFFF')
                            cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                            right=pyxl.styles.Side(border_style='thin'),
                            top=pyxl.styles.Side(border_style='thin'),
                            bottom=pyxl.styles.Side(border_style='thin'))
                        elif self.ee[s.assignee].totShiftHrs(s,styling=True)==12: #Colour 12 hrs shift green
                            cl.font=pyxl.styles.Font(bold=True,size=14)   
                            cl.alignment=pyxl.styles.Alignment(horizontal='center')
                            cl.fill=pyxl.styles.PatternFill(fill_type="solid",start_color='0092D050',end_color='0092D050')
                            cl.border=pyxl.styles.Border(left=pyxl.styles.Side(border_style='thin'),
                            right=pyxl.styles.Side(border_style='thin'),
                            top=pyxl.styles.Side(border_style='thin'),
                            bottom=pyxl.styles.Side(border_style='thin'))
                            
        #Initial Setup
        self.rev+=1
        wb=pyxl.Workbook()
        dest_filename = 'Wknd Sched_Rev '+str(self.rev)+'.xlsx'
        ws = wb.active
        ws.title = "Full Schedule" #Title worksheet for printout
        ws['A1']='Rev '+str(self.rev)
        styleCell(ws['A1'],'shift')
        ws['A2']='Manager'
        styleCell(ws['A2'],'shift')
        ws.column_dimensions['A'].width =22.78
        dys=['Friday','Saturday','Sunday','Monday']
        shifts=['C','A','B']*4
        tSlots=['11p - 3a','3a - 7a','7a - 11a', '11a - 3p','3p-7p', '7p-11p']*4
        for i in range(0,24,6): #Print Days of week
            cl=ws.cell(column=2+i,row=1)
            styleCell(cl,'day',6) #Style all the cells within the merge
            ws.merge_cells(start_row=1, start_column=2+i, end_row=1, end_column=2+i+5)
            cl.value=dys.pop(0)
        for i in range(0,24,2): #Print the shift title
            cl=ws.cell(column=2+i,row=3)
            styleCell(cl,'shift',1)
            ws.merge_cells(start_row=3, start_column=2+i, end_row=3, end_column=2+i+1)
            cl.value=shifts.pop(0)
        for i in range(24): #Print the shift times row
            cl=ws.cell(column=2+i,row=4)
            cl.value=tSlots.pop(0)
            styleCell(cl,'hours')

        def styleNfill(cl,s):
            if s.assnType=='DNS': cl.value="N/A"
            elif s.assignee is not None:
                val=self.ee[s.assignee].dispNm()
                if s.assnType=='F': val=val+'(F)' #Append force identifier
                cl.value=val
            else: 
                s.assnType='N'
                cl.value='NO STAFF'
            styleCell(cl,s.assnType,s)

        #==================
        #Enter shifts
        #The sequence of keys returned from self.slots.keys() is the same as the slots required were defined in "AllSlots"
        #They should all be in order which would allow for a naive method here but I will generalize it in case someone changes the template
        #And they are not in order after all.
        #The method here is to use a dictionary keyed by job name to track which row a job is printed out to,
        #and to add a job to the dictionary if/when it is encountered for the first time.
        jrD={} #job/Row Dict {job:row}
        r=5 #Start printing job slot info's at row 5 in excel
        for k in self.slots:
            s=self.slots[k] #Retrieve slot
            if s.dispNm not in jrD: #jobRowDict - Add to dict if first time seeing a job
                jrD[s.dispNm]=r
                jbCl=ws.cell(row=r,column=1)
                jbCl.value=s.dispNm
                styleCell(jbCl,'jbNm')
                r+=1 #increment for next one to be observed
            cl=ws.cell(row=jrD[s.dispNm],column=1+s.seqID)
            styleNfill(cl,s)
        #=========================================
        #Go through the schedule to add the (1/3),(1/2),(2/2) etc etc etc slot identifiers for human readibility
        #Define helper function
        def ranges(nums):
            """Returns a list of (open,close) intervals a list spans"""
            nums = sorted(set(nums))
            gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s+1 < e]
            edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
            return list(zip(edges, edges))
        #===================
        #Proceed with helper function
        for k in self.slots:
            s=self.slots[k] #Retrieve slot
            if s.assignee!=None:
                ids=sorted([self.slots[k].seqID for k in self.ee[s.assignee].assignments])
                rn=ranges(ids)#sets of start-end intervals
                myRn=[x for x in rn if s.seqID>= x[0] and s.seqID<= x[1]][0] #Retrieve start and end sqID for shift containing slot in question
                mysls=[self.slots[k] for k in self.ee[s.assignee].assignments if self.slots[k].seqID>=min(myRn) and self.slots[k].seqID<=max(myRn)] #Retrieve slots
                contents= [[s.dispNm,s.assnType] for s in mysls] #pull slots jobs/assn.Type
                denom=1
                for i in range(len(contents)-1):
                    if contents[i]!=contents[i+1]:
                        denom+=1
                numrtr=[x.key() for x in sorted(mysls, key=lambda x: x.seqID)].index(s.key())+1 #Check for index of current slot within myRn to see its numerator in sequence
                x=ws.cell(row=jrD[s.dispNm],column=1+s.seqID).value
                if denom!=1:
                    ws.cell(row=jrD[s.dispNm],column=1+s.seqID).value=str(x)+' ('+str(numrtr)+'/'+str(denom)+')'
        #======================================================
        #Merge contiguous shifts cells
        #- Define a custom function to facilitate it
        def numInARow(cl,n=1):
            """Given a cell, return the number of cells in a row have the same name in them... Forced shifts break the count. Count starts from 1 for a voluntary shift immediately following a forced shift"""
            nextval=cl.offset(0,1).value
            if nextval is None or nextval=='':
                return n
            elif cl.value!='' and cl.value is not None:
                if '/' in nextval:
                    nextval=cl.offset(0,1).value[:cl.offset(0,1).value.index('/')-3]
                curVal=cl.value
                if '/' in cl.value:
                    curVal=cl.value[:cl.value.index('/')-3]
                if nextval==curVal:
                    return numInARow(cl.offset(0,1),n+1) #Recursive fn. If next cell matches current, use this function on the next one again
                else: return n
            else: return n
        #====== Proceed with above function to be used
        nSkip=0 #Initialize for skipping cells in this loop as applicable
        for rw in range(5,r):
            for i in range(2,26):
                if nSkip<1:
                    inArow=numInARow(ws.cell(row=rw,column=i))
                    nSkip+=(inArow-1)
                    ws.merge_cells(start_row=rw, start_column=i, end_row=rw, end_column=i+inArow-1)
                elif nSkip>0: 
                    nSkip=nSkip-1 #This facilitates *not* checking cells that have already been merged.. thats because if i,i+1,i+2 merged at i, then when loop increments to i+1, it would give an error when checking on i+2.. need to skip to i+3 after having merged i,+1,+2
        #======================================
        #Format column widths
        #Doesnt appear to be working for some reason :(  
        for k in self.slots:
            s=self.slots[k] #Retrieve slot
            ws.column_dimensions[chr(65+s.seqID)].width = max(10.33,len(cl.value),ws.column_dimensions[chr(64+s.seqID)].width-5) #Widen column if new value is wider than any previously existing
        #=============================================
        #Print all forcings
        ws3 = wb.create_sheet(title="FORCINGS")
        ws3.cell(row=1,column=1).value='Employee ID'
        ws3.cell(row=1,column=2).value='Time slot'
        c=0
        for k in [k for k in self.slots if self.slots[k].assnType=='F']:
            ws3.cell(row=2+c,column=1).value=self.ee[self.slots[k].assignee].dispNm()
            ws3.cell(row=2+c,column=2).value=self.slots[k].dispNm+' '+ self.slLeg[self.slots[k].seqID-1][2]+' ('+self.slLeg[self.slots[k].seqID-1][1]+')'
            c+=1
        #=============================================
        #Print assignments to a separate sheet, sequenced by seniority
        ws2 = wb.create_sheet(title="Assignments (Sen'ty)")
        ws2.cell(row=2,column=1).value='Seniority'
        ws2.cell(row=2,column=2).value='Employee ID'
        ws2.cell(row=2,column=3).value='Time slots'
        # ws2.cell(row=1,column=1).value='Note that the seniority value presented is not actual plant seniority number, but just the sequence of '
        n=1
        for i in range(len(self.senList)-1):
            eId=self.senList[i][2]
            if len(self.ee[eId].assignments)>0 and self.slots[self.ee[eId].assignments[0]].assnType!='WWF':#If the person has an assignment and isn't WWF, print it
                n+=1
                ws2.cell(row=n+1,column=2).value=self.senList[i][0]
                ws2.cell(row=n+1,column=2).value=eId
                c=0
                for k in sorted(self.ee[eId].assignments,key=lambda k:int(k[:k.index('_')])):
                    styleNfill(ws2.cell(row=n+1,column=3+c),self.slots[k])
                    ws2.cell(row=n+1,column=3+c).value=self.slots[k].dispNm+' '+ self.slLeg[self.slots[k].seqID-1][2]+' ('+self.slLeg[self.slots[k].seqID-1][1]+')'
                    c+=1
        #=============================================
        #Print assignments to a separate sheet, in alphabetical order by last name
        ws2 = wb.create_sheet(title="Assignments (Alpha)")
        ws2.cell(row=2,column=1).value='Last, First'
        ws2.cell(row=2,column=2).value='Time slots'
        # ws2.cell(row=1,column=1).value='Note that the seniority value presented is not actual plant seniority number, but just the sequence of '
        n=1
        for rec in tls.viewTBL('senRef',sortBy=[('last','ASC')]):
            eId=rec[2]
            if len(self.ee[eId].assignments)>0 and self.slots[self.ee[eId].assignments[0]].assnType!='WWF':#If the person has an assignment and isn't WWF, print it
                n+=1
                ws2.cell(row=n+1,column=1).value=self.ee[eId].lastNm+', '+self.ee[eId].firstNm[0]+'.'
                # ws2.cell(row=n+1,column=2).value=eId
                c=0
                for k in sorted(self.ee[eId].assignments,key=lambda k:int(k[:k.index('_')])):
                    styleNfill(ws2.cell(row=n+1,column=2+c),self.slots[k])
                    # ws2.cell(row=n+1,column=2+c).value=self.slots[k].dispNm+' '+ self.slLeg[self.slots[k].seqID-1][2]+' ('+self.slLeg[self.slots[k].seqID-1][1]+')'
                    ws2.cell(row=n+1,column=2+c).value=self.slLeg[self.slots[k].seqID-1][2]+' ('+self.slLeg[self.slots[k].seqID-1][1]+')'
                    c+=1
        #==========================

        wb.save(filename = dest_filename)
        return dest_filename
        